from flask import Flask, request, jsonify, send_file, render_template
from flask_cors import CORS
import requests
import os
from dotenv import load_dotenv
import json
import threading
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import io
import pickle
from datetime import datetime
import math

# Загружаем переменные окружения
# ПРИМЕЧАНИЕ: Файл .env существует в проекте и содержит актуальные ключи API
load_dotenv()

app = Flask(__name__)
CORS(app)

# Глобальные переменные для управления процессами
current_processes = {
    'search_names': False,
    'search_emails': False
}

organizations_data = []

# Маршрут для главной страницы
@app.route('/')
def index():
    return render_template('index.html')

# Функции для работы с файловым хранилищем
def save_organizations_data(data, city):
    """Сохраняет данные организаций в файл"""
    try:
        filename = f"data_{city.replace(' ', '_')}.pkl"
        filepath = os.path.join('exports', filename)
        
        # Создаем директорию если не существует
        os.makedirs('exports', exist_ok=True)
        
        with open(filepath, 'wb') as f:
            pickle.dump(data, f)
        print(f"💾 Данные сохранены в файл: {filepath}")
        return filepath
    except Exception as e:
        print(f"❌ Ошибка сохранения данных: {e}")
        return None

def load_organizations_data(city):
    """Загружает данные организаций из файла"""
    try:
        filename = f"data_{city.replace(' ', '_')}.pkl"
        filepath = os.path.join('exports', filename)
        
        if os.path.exists(filepath):
            with open(filepath, 'rb') as f:
                data = pickle.load(f)
            print(f"📂 Данные загружены из файла: {filepath}, количество: {len(data)}")
            return data
        else:
            print(f"📂 Файл не найден: {filepath}")
            return []
    except Exception as e:
        print(f"❌ Ошибка загрузки данных: {e}")
        return []

class YandexSearchAPI:
    def __init__(self):
        # ПРИМЕЧАНИЕ: Файл .env существует в проекте и содержит актуальные ключи API
        self.api_key = os.getenv('YANDEX_SEARCH__API_KEY')
        self.base_url = 'https://search-maps.yandex.ru/v1/'
    
    def radius_to_spn(self, radius_km, latitude):
        """Преобразует радиус в километрах в параметр spn для Яндекс API"""
        # 1 градус широты ≈ 111 км
        lat_spn = radius_km / 111.0
        
        # 1 градус долготы зависит от широты
        # На экваторе: 1° ≈ 111 км, на полюсах: 1° ≈ 0 км
        lon_spn = radius_km / (111.0 * abs(math.cos(math.radians(latitude))))
        
        return f"{lon_spn:.6f},{lat_spn:.6f}"
        
    
    def search_organizations(self, city=None, selected_types=None, stop_flag=None, coordinates=None, radius=5):
        """Поиск курортных организаций в заданном городе или по координатам"""
        print(f"🔑 API ключ загружен: {'Да' if self.api_key else 'Нет'}")
        print(f"🔑 Выбранные типы: {selected_types}")
        
        # Определяем режим поиска
        search_by_coordinates = coordinates is not None and len(coordinates) == 2
        if search_by_coordinates:
            lon, lat = coordinates
            print(f"🎯 Режим поиска: по координатам {lat:.6f}, {lon:.6f}, радиус {radius} км")
        else:
            print(f"🏙️ Режим поиска: по названию города '{city}'")
        
        if not self.api_key:
            print("❌ API ключ не найден")
            return {'error': 'API ключ не найден'}
        
        if search_by_coordinates:
            print(f"Поиск организаций по координатам: {lat:.6f}, {lon:.6f}, радиус {radius} км")
        else:
            print(f"Поиск организаций в городе: {city}")
        
        # Используем только выбранные типы организаций
        organization_types = selected_types
        
        results = []
        
        for i, org_type in enumerate(organization_types):
            print(f"[{i+1}/{len(organization_types)}] Обрабатываем тип: {org_type}")
            
            if stop_flag and stop_flag():
                print(f"Процесс остановлен на типе: {org_type}")
                break
            
            # Добавляем небольшую задержку для демонстрации работы кнопки СТОП
            import time
            time.sleep(1)
                
            if search_by_coordinates:
                # Поиск по координатам и радиусу
                query = org_type
                print(f"Формируем запрос по координатам: '{query}'")
                print(f"🔍 Тип организации: '{org_type}'")
                print(f"🔍 Координаты: {lat:.6f}, {lon:.6f}")
                print(f"🔍 Радиус: {radius} км")
                
                # Преобразуем радиус в параметр spn для Яндекс API
                spn = self.radius_to_spn(radius, lat)
                print(f"🔍 SPN параметр: {spn}")
                
                params = {
                    'text': query,
                    'type': 'biz',
                    'lang': 'ru_RU',
                    'apikey': self.api_key,
                    'll': f"{lon},{lat}",  # Центр поиска
                    'spn': spn,  # Размер области поиска
                    'results': 20
                }
            else:
                # Поиск по названию города (старый алгоритм)
                query = f"{org_type} {city}"
                print(f"Формируем запрос: '{query}'")
                print(f"🔍 Тип организации: '{org_type}'")
                print(f"🔍 Город: '{city}'")
                params = {
                    'text': query,
                    'type': 'biz',
                    'lang': 'ru_RU',
                    'apikey': self.api_key,
                    'results': 20
                }
            
            try:
                print(f"Отправляем запрос к API: {self.base_url}")
                response = requests.get(self.base_url, params=params, timeout=10)
                print(f"Получен ответ: статус {response.status_code}")
                
                if response.status_code == 403:
                    print(f"❌ Превышен лимит запросов для типа '{org_type}'")
                    continue
                
                if response.status_code == 200:
                    data = response.json()
                    print(f"📊 Структура ответа Yandex API: {list(data.keys())}")
                    
                    features = data.get('features', [])
                    print(f"📊 API вернул {len(features)} объектов для типа '{org_type}'")
                    
                    # Показываем структуру первого объекта для анализа
                    if features:
                        first_feature = features[0]
                        print(f"📊 Структура объекта: {list(first_feature.keys())}")
                        print(f"📊 Properties: {list(first_feature.get('properties', {}).keys())}")
                        print(f"📊 Geometry: {list(first_feature.get('geometry', {}).keys())}")
                    
                    if len(features) == 0:
                        print(f"Нет организаций типа '{org_type}' в городе '{city}'")
                        continue
                    
                    added_count = 0
                    for j, feature in enumerate(features):
                        if stop_flag():
                            print(f"Процесс остановлен при обработке объекта {j+1} типа {org_type}")
                            break
                            
                        properties = feature.get('properties', {})
                        geometry = feature.get('geometry', {})
                        
                        org_name = properties.get('name', '')
                        org_description = properties.get('description', '')
                        
                        print(f"  [{j+1}/{len(features)}] Организация: '{org_name}'")
                        print(f"      Адрес: '{org_description}'")
                        
                        # Извлекаем данные из CompanyMetaData
                        company_meta = properties.get('CompanyMetaData', {})
                        yandex_id = company_meta.get('id', '')
                        full_address = company_meta.get('address', org_description)
                        website = company_meta.get('url', '')
                        
                        print(f"      ID из CompanyMetaData: {yandex_id}")
                        print(f"      Адрес из CompanyMetaData: {full_address}")
                        print(f"      Веб-сайт из CompanyMetaData: {website}")
                        
                        org_data = {
                            'name': org_name,
                            'coordinates': geometry.get('coordinates', []),
                            'yandex_id': yandex_id or f"yandex_{len(results)+1:04d}_{org_type.replace(' ', '_')}",
                            'full_address': full_address or org_description,
                            'website': website or f"https://{org_name[:15].replace(' ', '').lower()}.ru",
                            'email': '',         # Будет заполнен LLM
                            'type': org_type,
                            'city': city
                        }
                        
                        if org_data['name'] and org_data not in results:
                            results.append(org_data)
                            added_count += 1
                            print(f"      ✅ Добавлена в результаты (ID: {org_data['yandex_id']}, Сайт: {org_data['website']})")
                        else:
                            print(f"      ❌ Пропущена (дубликат или пустое название)")
                    
                    print(f"Добавлено {added_count} новых организаций типа '{org_type}'")
                else:
                    print(f"❌ Ошибка API: {response.status_code}")
                    print(f"Ответ сервера: {response.text[:200]}...")
                    
            except Exception as e:
                print(f"❌ Исключение при поиске {org_type}: {e}")
                continue
            
            # Добавляем задержку после обработки каждого типа
            time.sleep(0.5)
                
            print(f"Завершен поиск типа '{org_type}'. Всего найдено: {len(results)}")
            print("-" * 50)
        
        print(f"Всего найдено организаций: {len(results)}")
        
        # Фильтруем результаты по выбранным типам (как в тестовых данных)
        print(f"🔍 До фильтрации: {len(results)} организаций")
        print(f"🔍 Выбранные типы: {selected_types}")
        for i, org in enumerate(results[:5]):  # Показываем первые 5
            print(f"  [{i+1}] {org.get('name', 'Без названия')} - Тип: '{org.get('type', 'Нет')}'")
        
        filtered_results = [org for org in results if org['type'] in selected_types]
        print(f"🔍 После фильтрации по типам: {len(filtered_results)} организаций")
        
        # Удаляем дубликаты по адресу и сайту
        deduplicated_results = self.remove_duplicates(filtered_results)
        print(f"🔍 После удаления дубликатов: {len(deduplicated_results)} организаций")
        
        return {'organizations': deduplicated_results}
    
    def remove_duplicates(self, organizations):
        """Удаляет дубликаты организаций по адресу и сайту"""
        seen_addresses = set()
        seen_websites = set()
        unique_organizations = []
        
        for org in organizations:
            # Получаем адрес и сайт
            address = org.get('full_address', '').strip().lower()
            website = org.get('website', '').strip().lower()
            
            # Создаем ключ для проверки дубликатов
            address_key = address if address else f"no_address_{org.get('yandex_id', '')}"
            website_key = website if website else f"no_website_{org.get('yandex_id', '')}"
            
            # Проверяем, не встречались ли уже такие адрес или сайт
            is_duplicate = False
            
            if address and address in seen_addresses:
                print(f"🔄 Найден дубликат по адресу: {org.get('name', 'Без названия')} - {address}")
                is_duplicate = True
            
            if website and website in seen_websites:
                print(f"🔄 Найден дубликат по сайту: {org.get('name', 'Без названия')} - {website}")
                is_duplicate = True
            
            if not is_duplicate:
                unique_organizations.append(org)
                seen_addresses.add(address_key)
                seen_websites.add(website_key)
            else:
                print(f"❌ Удален дубликат: {org.get('name', 'Без названия')}")
        
        removed_count = len(organizations) - len(unique_organizations)
        if removed_count > 0:
            print(f"🧹 Удалено дубликатов: {removed_count}")
        
        return unique_organizations
    
    def get_organization_details_by_coordinates(self, lon, lat, stop_flag):
        """Получение детальной информации об организации по координатам"""
        if not self.api_key:
            return {'error': 'API ключ не найден'}
        
        # Используем обратный геокодинг для получения информации по координатам
        params = {
            'geocode': f"{lon},{lat}",
            'kind': 'house',
            'format': 'json',
            'results': 1,
            'lang': 'ru_RU',
            'apikey': self.api_key
        }
        
        try:
            print(f"      🔍 Запрашиваем детали по координатам: {lat}, {lon}")
            response = requests.get("https://geocode-maps.yandex.ru/1.x/", params=params, timeout=10)
            
            if response.status_code == 200:
                data = response.json()
                features = data.get('response', {}).get('GeoObjectCollection', {}).get('featureMember', [])
                
                if features:
                    feature = features[0]
                    properties = feature.get('GeoObject', {}).get('metaDataProperty', {}).get('GeocoderMetaData', {})
                    
                    # Извлекаем информацию
                    yandex_id = properties.get('id', '')
                    full_address = properties.get('text', '')
                    
                    print(f"      📍 Найдена информация: ID={yandex_id[:20]}..., Адрес={full_address[:50]}...")
                    
                    return {
                        'yandex_id': yandex_id,
                        'full_address': full_address,
                        'website': '',  # В геокодинге нет веб-сайта
                        'phone': '',
                        'hours': ''
                    }
                else:
                    print(f"      ❌ Нет данных по координатам")
                    return {'error': 'Нет данных по координатам'}
            else:
                print(f"      ❌ Ошибка геокодинга: {response.status_code}")
                return {'error': f'Ошибка геокодинга: {response.status_code}'}
                
        except Exception as e:
            print(f"      ❌ Исключение при геокодинге: {e}")
            return {'error': f'Ошибка запроса: {e}'}

    def search_website_by_name(self, org_name, city, stop_flag):
        """Поиск веб-сайта организации по названию"""
        if not self.api_key:
            return {'error': 'API ключ не найден'}
        
        # Ищем организацию по точному названию
        params = {
            'text': f"{org_name} {city}",
            'type': 'biz',
            'lang': 'ru_RU',
            'results': 1,
            'apikey': self.api_key
        }
        
        try:
            print(f"      🌐 Ищем веб-сайт для: {org_name}")
            response = requests.get(self.base_url, params=params, timeout=10)
            
            if response.status_code == 200:
                data = response.json()
                features = data.get('features', [])
                
                if features:
                    feature = features[0]
                    properties = feature.get('properties', {})
                    
                    # Проверяем различные поля на наличие веб-сайта
                    website = (properties.get('website', '') or 
                             properties.get('url', '') or 
                             properties.get('site', '') or 
                             properties.get('web', ''))
                    
                    if website:
                        print(f"      🌐 Найден веб-сайт: {website}")
                        return {'website': website}
                    else:
                        print(f"      ❌ Веб-сайт не найден")
                        return {'error': 'Веб-сайт не найден'}
                else:
                    print(f"      ❌ Организация не найдена для поиска веб-сайта")
                    return {'error': 'Организация не найдена'}
            else:
                print(f"      ❌ Ошибка поиска веб-сайта: {response.status_code}")
                return {'error': f'Ошибка поиска: {response.status_code}'}
                
        except Exception as e:
            print(f"      ❌ Исключение при поиске веб-сайта: {e}")
            return {'error': f'Ошибка запроса: {e}'}

    def get_organization_details(self, yandex_id, stop_flag):
        """Получение детальной информации об организации по ID"""
        if not self.api_key or not yandex_id:
            return {'error': 'Недостаточно данных для поиска'}
        
        params = {
            'id': yandex_id,
            'lang': 'ru_RU',
            'apikey': self.api_key
        }
        
        try:
            response = requests.get(f"{self.base_url}details", params=params, timeout=10)
            if response.status_code == 200:
                data = response.json()
                properties = data.get('properties', {})
                
                # Извлекаем полный адрес
                address_parts = []
                if properties.get('address'):
                    address_parts.append(properties['address'])
                if properties.get('description'):
                    address_parts.append(properties['description'])
                
                full_address = ', '.join(filter(None, address_parts))
                
                return {
                    'full_address': full_address,
                    'website': properties.get('website', ''),
                    'phone': properties.get('phone', ''),
                    'hours': properties.get('hours', '')
                }
            else:
                return {'error': f'Ошибка API: {response.status_code}'}
        except Exception as e:
            return {'error': f'Ошибка запроса: {e}'}

class ProxyAPIClient:
    def __init__(self):
        self.api_key = os.getenv('PROXYAPI_KEY')
        self.base_url = os.getenv('PROXYAPI_BASE_URL')
    
    def search_email(self, organization_name, city, stop_flag):
        """Поиск email организации через LLM"""
        if not self.api_key or not self.base_url:
            return {'error': 'ProxyAPI credentials не найдены'}
        
        prompt = f"""
        Найди официальный email адрес для организации "{organization_name}" в городе {city}.
        Организация предоставляет услуги размещения отдыхающих (база отдыха, дом отдыха, гостиница, санаторий, гостевой дом, хостел).
        
        Верни только email адрес, если найдешь. Если не найдешь, верни "не найден".
        """
        
        try:
            headers = {
                'Authorization': f'Bearer {self.api_key}',
                'Content-Type': 'application/json'
            }
            
            data = {
                'model': 'gpt-3.5-turbo',
                'messages': [
                    {'role': 'user', 'content': prompt}
                ],
                'max_tokens': 100
            }
            
            response = requests.post(
                f"{self.base_url}/v1/chat/completions",
                headers=headers,
                json=data,
                timeout=30
            )
            
            if response.status_code == 200:
                result = response.json()
                email = result.get('choices', [{}])[0].get('message', {}).get('content', '').strip()
                if '@' in email and email != 'не найден':
                    return {'email': email}
                else:
                    return {'email': 'не найден'}
            else:
                return {'error': f'Ошибка API: {response.status_code}'}
                
        except Exception as e:
            return {'error': f'Ошибка запроса: {e}'}

# Инициализация API клиентов
yandex_api = YandexSearchAPI()
proxy_api = ProxyAPIClient()

# Главная страница теперь обслуживается фронтендом

def determine_region_by_coordinates(lat, lon):
    """Определяет регион по координатам"""
    # Основные регионы России с примерными границами
    regions = {
        'Крым': {'lat_min': 44.0, 'lat_max': 46.0, 'lon_min': 32.0, 'lon_max': 37.0},
        'Краснодарский край': {'lat_min': 43.0, 'lat_max': 47.0, 'lon_min': 36.0, 'lon_max': 42.0},
        'Московская область': {'lat_min': 54.0, 'lat_max': 57.0, 'lon_min': 35.0, 'lon_max': 40.0},
        'Санкт-Петербург': {'lat_min': 59.0, 'lat_max': 61.0, 'lon_min': 29.0, 'lon_max': 31.0},
        'Ленинградская область': {'lat_min': 58.0, 'lat_max': 61.0, 'lon_min': 28.0, 'lon_max': 35.0},
        'Новосибирская область': {'lat_min': 53.0, 'lat_max': 57.0, 'lon_min': 75.0, 'lon_max': 85.0},
        'Свердловская область': {'lat_min': 56.0, 'lat_max': 61.0, 'lon_min': 57.0, 'lon_max': 66.0},
        'Челябинская область': {'lat_min': 52.0, 'lat_max': 57.0, 'lon_min': 57.0, 'lon_max': 65.0},
        'Ростовская область': {'lat_min': 46.0, 'lat_max': 50.0, 'lon_min': 38.0, 'lon_max': 45.0},
        'Волгоградская область': {'lat_min': 48.0, 'lat_max': 51.0, 'lon_min': 41.0, 'lon_max': 48.0},
        'Ставропольский край': {'lat_min': 44.0, 'lat_max': 46.0, 'lon_min': 40.0, 'lon_max': 46.0},
        'Приморский край': {'lat_min': 42.0, 'lat_max': 48.0, 'lon_min': 130.0, 'lon_max': 140.0},
        'Хабаровский край': {'lat_min': 46.0, 'lat_max': 55.0, 'lon_min': 130.0, 'lon_max': 145.0},
        'Иркутская область': {'lat_min': 51.0, 'lat_max': 60.0, 'lon_min': 95.0, 'lon_max': 120.0},
        'Красноярский край': {'lat_min': 51.0, 'lat_max': 70.0, 'lon_min': 80.0, 'lon_max': 110.0},
        'Тюменская область': {'lat_min': 55.0, 'lat_max': 70.0, 'lon_min': 60.0, 'lon_max': 80.0},
        'Омская область': {'lat_min': 53.0, 'lat_max': 58.0, 'lon_min': 70.0, 'lon_max': 78.0},
        'Томская область': {'lat_min': 55.0, 'lat_max': 61.0, 'lon_min': 75.0, 'lon_max': 90.0},
        'Кемеровская область': {'lat_min': 52.0, 'lat_max': 57.0, 'lon_min': 84.0, 'lon_max': 90.0},
        'Алтайский край': {'lat_min': 50.0, 'lat_max': 54.0, 'lon_min': 78.0, 'lon_max': 87.0},
        'Республика Башкортостан': {'lat_min': 51.0, 'lat_max': 56.0, 'lon_min': 53.0, 'lon_max': 60.0},
        'Республика Татарстан': {'lat_min': 54.0, 'lat_max': 57.0, 'lon_min': 47.0, 'lon_max': 54.0},
        'Нижегородская область': {'lat_min': 54.0, 'lat_max': 58.0, 'lon_min': 40.0, 'lon_max': 48.0},
        'Самарская область': {'lat_min': 51.0, 'lat_max': 55.0, 'lon_min': 48.0, 'lon_max': 54.0},
        'Саратовская область': {'lat_min': 49.0, 'lat_max': 53.0, 'lon_min': 42.0, 'lon_max': 50.0},
        'Воронежская область': {'lat_min': 49.0, 'lat_max': 52.0, 'lon_min': 38.0, 'lon_max': 43.0},
        'Белгородская область': {'lat_min': 49.0, 'lat_max': 51.0, 'lon_min': 35.0, 'lon_max': 39.0},
        'Курская область': {'lat_min': 50.0, 'lat_max': 52.0, 'lon_min': 34.0, 'lon_max': 38.0},
        'Липецкая область': {'lat_min': 51.0, 'lat_max': 53.0, 'lon_min': 37.0, 'lon_max': 40.0},
        'Тамбовская область': {'lat_min': 51.0, 'lat_max': 54.0, 'lon_min': 40.0, 'lon_max': 43.0},
        'Рязанская область': {'lat_min': 53.0, 'lat_max': 55.0, 'lon_min': 38.0, 'lon_max': 42.0},
        'Тульская область': {'lat_min': 53.0, 'lat_max': 55.0, 'lon_min': 35.0, 'lon_max': 39.0},
        'Калужская область': {'lat_min': 53.0, 'lat_max': 55.0, 'lon_min': 33.0, 'lon_max': 37.0},
        'Смоленская область': {'lat_min': 53.0, 'lat_max': 56.0, 'lon_min': 30.0, 'lon_max': 35.0},
        'Брянская область': {'lat_min': 52.0, 'lat_max': 54.0, 'lon_min': 31.0, 'lon_max': 35.0},
        'Орловская область': {'lat_min': 52.0, 'lat_max': 54.0, 'lon_min': 35.0, 'lon_max': 38.0},
        'Костромская область': {'lat_min': 57.0, 'lat_max': 59.0, 'lon_min': 40.0, 'lon_max': 45.0},
        'Ивановская область': {'lat_min': 56.0, 'lat_max': 58.0, 'lon_min': 39.0, 'lon_max': 43.0},
        'Владимирская область': {'lat_min': 55.0, 'lat_max': 57.0, 'lon_min': 38.0, 'lon_max': 42.0},
        'Ярославская область': {'lat_min': 56.0, 'lat_max': 59.0, 'lon_min': 37.0, 'lon_max': 42.0},
        'Тверская область': {'lat_min': 55.0, 'lat_max': 58.0, 'lon_min': 31.0, 'lon_max': 37.0},
        'Калининградская область': {'lat_min': 54.0, 'lat_max': 55.0, 'lon_min': 19.0, 'lon_max': 20.0},
        'Архангельская область': {'lat_min': 60.0, 'lat_max': 67.0, 'lon_min': 35.0, 'lon_max': 50.0},
        'Вологодская область': {'lat_min': 58.0, 'lat_max': 61.0, 'lon_min': 35.0, 'lon_max': 42.0},
        'Мурманская область': {'lat_min': 66.0, 'lat_max': 70.0, 'lon_min': 28.0, 'lon_max': 40.0},
        'Республика Карелия': {'lat_min': 60.0, 'lat_max': 67.0, 'lon_min': 29.0, 'lon_max': 37.0},
        'Республика Коми': {'lat_min': 59.0, 'lat_max': 68.0, 'lon_min': 45.0, 'lon_max': 66.0},
        'Ненецкий автономный округ': {'lat_min': 66.0, 'lat_max': 70.0, 'lon_min': 43.0, 'lon_max': 60.0},
        'Ханты-Мансийский автономный округ': {'lat_min': 58.0, 'lat_max': 66.0, 'lon_min': 60.0, 'lon_max': 85.0},
        'Ямало-Ненецкий автономный округ': {'lat_min': 64.0, 'lat_max': 73.0, 'lon_min': 60.0, 'lon_max': 80.0},
        'Чукотский автономный округ': {'lat_min': 62.0, 'lat_max': 72.0, 'lon_min': 160.0, 'lon_max': 180.0},
        'Магаданская область': {'lat_min': 58.0, 'lat_max': 66.0, 'lon_min': 140.0, 'lon_max': 160.0},
        'Сахалинская область': {'lat_min': 45.0, 'lat_max': 55.0, 'lon_min': 140.0, 'lon_max': 150.0},
        'Камчатский край': {'lat_min': 50.0, 'lat_max': 65.0, 'lon_min': 155.0, 'lon_max': 165.0},
        'Амурская область': {'lat_min': 48.0, 'lat_max': 57.0, 'lon_min': 119.0, 'lon_max': 135.0},
        'Еврейская автономная область': {'lat_min': 47.0, 'lat_max': 49.0, 'lon_min': 130.0, 'lon_max': 135.0},
        'Забайкальский край': {'lat_min': 49.0, 'lat_max': 58.0, 'lon_min': 107.0, 'lon_max': 120.0},
        'Республика Бурятия': {'lat_min': 50.0, 'lat_max': 58.0, 'lon_min': 98.0, 'lon_max': 116.0},
        'Республика Саха (Якутия)': {'lat_min': 55.0, 'lat_max': 75.0, 'lon_min': 105.0, 'lon_max': 170.0},
        'Республика Тыва': {'lat_min': 49.0, 'lat_max': 54.0, 'lon_min': 88.0, 'lon_max': 100.0},
        'Республика Хакасия': {'lat_min': 51.0, 'lat_max': 55.0, 'lon_min': 87.0, 'lon_max': 92.0},
        'Республика Алтай': {'lat_min': 49.0, 'lat_max': 52.0, 'lon_min': 84.0, 'lon_max': 90.0},
        'Республика Калмыкия': {'lat_min': 45.0, 'lat_max': 48.0, 'lon_min': 44.0, 'lon_max': 47.0},
        'Республика Адыгея': {'lat_min': 44.0, 'lat_max': 45.0, 'lon_min': 38.0, 'lon_max': 40.0},
        'Республика Карачаево-Черкесия': {'lat_min': 43.0, 'lat_max': 45.0, 'lon_min': 40.0, 'lon_max': 42.0},
        'Республика Кабардино-Балкария': {'lat_min': 43.0, 'lat_max': 44.0, 'lon_min': 42.0, 'lon_max': 44.0},
        'Республика Северная Осетия': {'lat_min': 42.0, 'lat_max': 44.0, 'lon_min': 43.0, 'lon_max': 45.0},
        'Республика Ингушетия': {'lat_min': 42.0, 'lat_max': 43.0, 'lon_min': 44.0, 'lon_max': 45.0},
        'Чеченская Республика': {'lat_min': 42.0, 'lat_max': 44.0, 'lon_min': 44.0, 'lon_max': 47.0},
        'Республика Дагестан': {'lat_min': 41.0, 'lat_max': 45.0, 'lon_min': 45.0, 'lon_max': 48.0},
        'Республика Мордовия': {'lat_min': 53.0, 'lat_max': 55.0, 'lon_min': 42.0, 'lon_max': 46.0},
        'Чувашская Республика': {'lat_min': 54.0, 'lat_max': 56.0, 'lon_min': 45.0, 'lon_max': 48.0},
        'Удмуртская Республика': {'lat_min': 55.0, 'lat_max': 58.0, 'lon_min': 51.0, 'lon_max': 54.0},
        'Пермский край': {'lat_min': 56.0, 'lat_max': 61.0, 'lon_min': 51.0, 'lon_max': 60.0},
        'Кировская область': {'lat_min': 56.0, 'lat_max': 61.0, 'lon_min': 45.0, 'lon_max': 52.0},
        'Республика Марий Эл': {'lat_min': 55.0, 'lat_max': 57.0, 'lon_min': 45.0, 'lon_max': 48.0},
        'Пензенская область': {'lat_min': 52.0, 'lat_max': 54.0, 'lon_min': 42.0, 'lon_max': 46.0},
        'Ульяновская область': {'lat_min': 52.0, 'lat_max': 55.0, 'lon_min': 46.0, 'lon_max': 50.0},
    }
    
    # Проверяем каждый регион
    for region_name, bounds in regions.items():
        if (bounds['lat_min'] <= lat <= bounds['lat_max'] and 
            bounds['lon_min'] <= lon <= bounds['lon_max']):
            return region_name
    
    # Если не найден, возвращаем общий регион
    return 'Россия'

def search_cities_2gis(city_name):
    """Поиск городов через 2GIS API"""
    print(f"🗺️ Используем 2GIS API для поиска: '{city_name}'")
    
    # Получаем API ключ 2GIS
    api_key = os.getenv('2GIS_API_KEY')
    if not api_key:
        print("❌ API ключ 2GIS не найден")
        return jsonify({'error': 'API ключ 2GIS не настроен'}), 500
    
    # 2GIS API для поиска городов
    search_url = 'https://catalog.api.2gis.com/3.0/items'
    
    params = {
        'q': city_name,
        'key': api_key,
        'type': 'adm_div',  # Используем правильный тип для административных единиц
        'fields': 'items.point,items.name,items.full_name,items.type,items.address',
        'page_size': 10
    }
    
    try:
        response = requests.get(search_url, params=params, timeout=10)
        print(f"📡 2GIS ответ: статус {response.status_code}")
        
        if response.status_code == 200:
            data = response.json()
            print(f"📊 2GIS нашел {len(data.get('result', {}).get('items', []))} объектов")
            
            cities = []
            items = data.get('result', {}).get('items', [])
            
            for i, item in enumerate(items):
                name = item.get('name', '')
                full_name = item.get('full_name', '')
                item_type = item.get('type', '')
                point = item.get('point', {})
                
                # Извлекаем координаты
                if point and 'lon' in point and 'lat' in point:
                    lon = float(point['lon'])
                    lat = float(point['lat'])
                else:
                    print(f"  [{i+1}] {name} - НЕТ КООРДИНАТ (пропускаем)")
                    continue
                
                # Фильтруем только административные единицы (города, районы)
                if 'adm_div' in item_type:
                    # Извлекаем регион из full_name или определяем по координатам
                    region = 'Россия'
                    country = 'Россия'
                    
                    # Улучшенная логика определения региона
                    if full_name and full_name != name:
                        import re
                        # Ищем регион в полном названии
                        region_patterns = [
                            r'([А-Яа-яё\s]+(?:край|область|республика|округ))',
                            r'([А-Яа-яё\s]+(?:федеральный округ))',
                            r'(Крым)',
                            r'(Севастополь)'
                        ]
                        
                        for pattern in region_patterns:
                            region_match = re.search(pattern, full_name)
                            if region_match:
                                region = region_match.group(1).strip()
                                break
                    
                    # Если регион не найден в названии, определяем по координатам
                    if region == 'Россия':
                        region = determine_region_by_coordinates(lat, lon)
                    
                    city_info = {
                        'name': name,
                        'coordinates': [lon, lat],
                        'region': region,
                        'country': country,
                        'full_address': full_name,
                        'search_type': '2gis'
                    }
                    cities.append(city_info)
                    print(f"  [{i+1}] {name} ({item_type}) - ГОРОД")
                    print(f"      Полное название: {full_name}")
                    print(f"      Регион: {region}")
                    print(f"      Координаты: {lat}, {lon}")
                else:
                    print(f"  [{i+1}] {name} ({item_type}) - НЕ ГОРОД (пропускаем)")
            
            if cities:
                print(f"✅ 2GIS нашел {len(cities)} городов")
                return jsonify({'cities': cities})
            else:
                print("❌ 2GIS не нашел городов")
                return jsonify({'error': 'Проверьте правильность написания названия города'}), 404
                
        else:
            print(f"❌ Ошибка 2GIS: {response.status_code}")
            print(f"Ответ сервера: {response.text[:200]}...")
            return jsonify({'error': 'Ошибка при обращении к 2GIS API'}), 500
            
    except Exception as e:
        print(f"❌ Исключение 2GIS: {e}")
        return jsonify({'error': 'Ошибка при поиске городов'}), 500

# def search_cities_nominatim(city_name):
#     """Поиск городов через Nominatim API (fallback) - ЗАКОММЕНТИРОВАНО"""
#     print(f"🌍 Используем Nominatim API для поиска: '{city_name}'")
#     
#     # Nominatim API для поиска городов
#     search_url = 'https://nominatim.openstreetmap.org/search'
#     
#     params = {
#         'q': city_name,
#         'format': 'json',
#         'countrycodes': 'ru',  # Только Россия
#         'limit': 5,
#         'addressdetails': 1
#     }
#     
#     headers = {
#         'User-Agent': 'ResortSearchApp/1.0 (contact@example.com)'
#     }
#     
#     try:
#         response = requests.get(search_url, params=params, headers=headers, timeout=10)
#         print(f"📡 Nominatim ответ: статус {response.status_code}")
#         
#         if response.status_code == 200:
#             data = response.json()
#             print(f"📊 Nominatim нашел {len(data)} объектов")
#             
#             cities = []
#             for i, item in enumerate(data):
#                 name = item.get('display_name', '')
#                 lat = float(item.get('lat', 0))
#                 lon = float(item.get('lon', 0))
#                 place_type = item.get('type', '')
#                 
#                 # Фильтруем только населенные пункты
#                 if place_type in ['village', 'city', 'town', 'hamlet', 'locality']:
#                     # Извлекаем регион из display_name
#                     region = 'Россия'
#                     country = 'Россия'
#                     
#                     if name:
#                         import re
#                         region_match = re.search(r'([А-Яа-яё\s]+(?:край|область|республика|округ))', name)
#                         if region_match:
#                             region = region_match.group(1).strip()
#                     
#                     city_info = {
#                         'name': name.split(',')[0] if ',' in name else name,  # Берем только название города
#                         'coordinates': [lon, lat],
#                         'region': region,
#                         'country': country,
#                         'full_address': name,
#                         'search_type': 'nominatim'
#                     }
#                     cities.append(city_info)
#                     print(f"  [{i+1}] {name} ({place_type}) - ГОРОД")
#                     print(f"      Координаты: {lat}, {lon}")
#                 else:
#                     print(f"  [{i+1}] {name} ({place_type}) - НЕ ГОРОД (пропускаем)")
#             
#             if cities:
#                 print(f"✅ Nominatim нашел {len(cities)} городов")
#                 return jsonify({'cities': cities})
#             else:
#                 print("❌ Nominatim не нашел городов")
#                 return jsonify({'error': 'Проверьте правильность написания названия города'}), 404
#                 
#         else:
#             print(f"❌ Ошибка Nominatim: {response.status_code}")
#             return jsonify({'error': 'Ошибка при обращении к Nominatim API'}), 500
#             
#     except Exception as e:
#         print(f"❌ Исключение Nominatim: {e}")
#         return jsonify({'error': 'Ошибка при поиске городов'}), 500

@app.route('/api/search_cities', methods=['POST'])
def search_cities():
    """Поиск городов через 2GIS API"""
    print(f"🏙️ Получен запрос на поиск городов")
    
    data = request.json
    city_name = data.get('city', '').strip()
    
    print(f"🔍 Ищем город: '{city_name}'")
    
    if not city_name:
        print("❌ Ошибка: Название города не указано")
        return jsonify({'error': 'Название города не указано'}), 400
    
    # Используем 2GIS API для поиска городов
    return search_cities_2gis(city_name)

@app.route('/api/search_organizations', methods=['POST'])
def search_organizations():
    global organizations_data, current_processes
    
    print(f"🚀 Получен запрос на поиск организаций")
    data = request.json
    
    # Поддерживаем как старый формат (по названию города), так и новый (по координатам)
    city = data.get('city', '').strip()
    coordinates = data.get('coordinates')  # [lon, lat]
    radius = data.get('radius', 5)  # Радиус в км, по умолчанию 5
    selected_types = data.get('types', [])
    
    print(f"🏙️ Город: '{city}'")
    print(f"📍 Координаты: {coordinates}")
    print(f"📏 Радиус: {radius} км")
    print(f"📋 Выбранные типы: {selected_types}")
    print(f"📊 Количество типов: {len(selected_types)}")
    
    # Проверяем параметры
    if not coordinates and not city:
        print("❌ Ошибка: Не указаны ни город, ни координаты")
        return jsonify({'error': 'Не указаны ни город, ни координаты'}), 400
    
    if not selected_types:
        print("❌ Ошибка: Не выбраны типы организаций")
        return jsonify({'error': 'Не выбраны типы организаций'}), 400
    
    # Если переданы координаты, используем новый алгоритм
    if coordinates and len(coordinates) == 2:
        lon, lat = coordinates
        print(f"🎯 Используем поиск по координатам: {lat:.6f}, {lon:.6f}, радиус {radius} км")
    else:
        print(f"🏙️ Используем поиск по названию города: {city}")
        if not city:
            print("❌ Ошибка: Город не указан")
            return jsonify({'error': 'Город не указан'}), 400
    
    # Сброс данных при каждом поиске
    global organizations_data
    organizations_data = []
    
    current_processes['search_names'] = True
    
    def search_task():
        global organizations_data, current_processes
        try:
            print(f"🚀 Запуск поиска организаций в городе: {city}")
            print(f"Флаг остановки: {current_processes['search_names']}")
            
            # Передаем параметры в зависимости от режима поиска
            if coordinates and len(coordinates) == 2:
                result = yandex_api.search_organizations(
                    city=None, 
                    selected_types=selected_types, 
                    stop_flag=lambda: not current_processes['search_names'],
                    coordinates=coordinates,
                    radius=radius
                )
            else:
                result = yandex_api.search_organizations(
                    city=city, 
                    selected_types=selected_types, 
                    stop_flag=lambda: not current_processes['search_names']
                )
            
            print(f"🔍 Результат поиска: {result}")
            
            if 'error' not in result:
                global organizations_data
                organizations_data = result['organizations']
                print(f"✅ Поиск завершен. Найдено {len(organizations_data)} организаций")
                print(f"📊 Данные сохранены в organizations_data: {len(organizations_data)} элементов")
                print(f"🔍 Проверка: organizations_data содержит {len(organizations_data)} элементов")
                
                # Сохраняем данные в файл для экспорта
                if coordinates and len(coordinates) == 2:
                    # Для поиска по координатам используем специальное имя файла
                    city_name_for_file = f"coords_{coordinates[1]:.4f}_{coordinates[0]:.4f}_r{radius}"
                    # Удаляем старый файл если существует
                    old_filepath = os.path.join('exports', f"data_{city_name_for_file.replace(' ', '_')}.pkl")
                    if os.path.exists(old_filepath):
                        os.remove(old_filepath)
                        print(f"🗑️ Удален старый файл: {old_filepath}")
                    save_organizations_data(organizations_data, city_name_for_file)
                else:
                    # Удаляем старый файл если существует
                    old_filepath = os.path.join('exports', f"data_{city.replace(' ', '_')}.pkl")
                    if os.path.exists(old_filepath):
                        os.remove(old_filepath)
                        print(f"🗑️ Удален старый файл: {old_filepath}")
                    save_organizations_data(organizations_data, city)
            else:
                print(f"❌ Ошибка поиска: {result['error']}")
                
        except Exception as e:
            print(f"❌ Исключение в поиске организаций: {e}")
        finally:
            current_processes['search_names'] = False
            print(f"🏁 Процесс поиска названий завершен. Флаг: {current_processes['search_names']}")
    
    # Запуск в отдельном потоке
    thread = threading.Thread(target=search_task)
    thread.start()
    
    return jsonify({'message': f'Поиск организаций в городе {city} запущен'})

@app.route('/api/search_emails', methods=['POST'])
def search_emails():
    global organizations_data, current_processes
    
    # Получаем город из запроса для загрузки данных
    data = request.json
    city = data.get('city', '').strip()
    
    # Загружаем данные из файла
    organizations_data = load_organizations_data(city)
    
    if not organizations_data:
        return jsonify({'error': 'Сначала найдите организации'}), 400
    
    current_processes['search_emails'] = True
    
    def email_search_task():
        global organizations_data, current_processes
        try:
            for i, org in enumerate(organizations_data):
                if not current_processes['search_emails']:
                    break
                
                if not org.get('email'):
                    result = proxy_api.search_email(org['name'], org.get('city', ''), 
                                                  lambda: not current_processes['search_emails'])
                    if 'email' in result:
                        organizations_data[i]['email'] = result['email']
                
                time.sleep(1)  # Задержка между запросами
        except Exception as e:
            print(f"Ошибка поиска email: {e}")
        finally:
            current_processes['search_emails'] = False
    
    thread = threading.Thread(target=email_search_task)
    thread.start()
    
    return jsonify({'message': 'Поиск email адресов запущен'})


@app.route('/api/get_organizations', methods=['GET'])
def get_organizations():
    global organizations_data
    
    # Получаем параметры из запроса
    city = request.args.get('city', '').strip()
    coordinates = request.args.get('coordinates', '').strip()
    radius = request.args.get('radius', '5').strip()
    
    # Сначала проверяем глобальные данные
    if organizations_data:
        print(f"📤 Возвращаем данные из глобальной переменной: {len(organizations_data)} организаций")
        data_to_return = organizations_data
    elif city:
        # Загружаем данные из файла по названию города
        data_to_return = load_organizations_data(city)
        print(f"📤 Запрос на получение организаций для города '{city}'. Загружено: {len(data_to_return)} организаций")
    elif coordinates:
        # Загружаем данные из файла по координатам
        try:
            coords_parts = coordinates.split(',')
            if len(coords_parts) == 2:
                lat, lon = float(coords_parts[0]), float(coords_parts[1])
                city_name_for_file = f"coords_{lat:.4f}_{lon:.4f}_r{radius}"
                data_to_return = load_organizations_data(city_name_for_file)
                print(f"📤 Запрос на получение организаций для координат '{coordinates}', радиус {radius} км. Загружено: {len(data_to_return)} организаций")
            else:
                data_to_return = []
                print(f"❌ Неверный формат координат: {coordinates}")
        except ValueError:
            data_to_return = []
            print(f"❌ Ошибка парсинга координат: {coordinates}")
    else:
        # Если ни город, ни координаты не указаны, возвращаем пустой список
        data_to_return = []
        print(f"📤 Запрос на получение организаций без указания города или координат. Возвращаем пустой список.")
    
    print(f"📊 Текущие процессы: {current_processes}")
    
    if data_to_return:
        print(f"📤 Отправляем {len(data_to_return)} организаций")
        for i, org in enumerate(data_to_return[:5]):  # Показываем только первые 5 для краткости
            print(f"  [{i+1}] {org.get('name', 'Без названия')} - ID: {org.get('yandex_id', 'Нет')} - Тип: {org.get('type', 'Нет')}")
        if len(data_to_return) > 5:
            print(f"  ... и еще {len(data_to_return) - 5} организаций")
    else:
        print("⚠️ Данные не найдены!")
        
    return jsonify({'organizations': data_to_return})

@app.route('/api/stop_process', methods=['POST'])
def stop_process():
    data = request.json
    process_type = data.get('process_type')
    
    print(f"🛑 Запрос на остановку процесса: {process_type}")
    print(f"Текущие процессы: {current_processes}")
    
    if process_type in current_processes:
        old_value = current_processes[process_type]
        current_processes[process_type] = False
        print(f"✅ Процесс {process_type} остановлен (было: {old_value}, стало: {current_processes[process_type]})")
        print(f"Обновленные процессы: {current_processes}")
        return jsonify({'message': f'Процесс {process_type} остановлен'})
    
    print(f"❌ Неизвестный тип процесса: {process_type}")
    print(f"Доступные процессы: {list(current_processes.keys())}")
    return jsonify({'error': 'Неизвестный тип процесса'}), 400

@app.route('/api/export_excel', methods=['GET'])
def export_excel():
    """Экспорт данных организаций в Excel файл"""
    try:
        # Получаем параметры из запроса
        city_name = request.args.get('city', '').strip()
        coordinates = request.args.get('coordinates', '').strip()
        radius = request.args.get('radius', '5').strip()
        
        print(f"📊 Параметры экспорта: city='{city_name}', coordinates='{coordinates}', radius='{radius}'")
        
        # Определяем источник данных
        if city_name and not coordinates:
            # Экспорт по названию города (старый алгоритм)
            data_to_export = load_organizations_data(city_name)
            file_source = f"город '{city_name}'"
        elif coordinates:
            # Экспорт по координатам
            try:
                coords_parts = coordinates.split(',')
                if len(coords_parts) == 2:
                    lat, lon = float(coords_parts[0]), float(coords_parts[1])
                    city_name_for_file = f"coords_{lat:.4f}_{lon:.4f}_r{radius}"
                    data_to_export = load_organizations_data(city_name_for_file)
                    
                    # Используем переданное название города или извлекаем из данных
                    if city_name:
                        # Используем название города, переданное из frontend
                        extracted_city_name = city_name
                        file_source = f"город '{city_name}' (по координатам)"
                        print(f"✅ Используем переданное название города: '{city_name}'")
                    else:
                        # Fallback: извлекаем название города из данных организаций
                        extracted_city_name = None
                        if data_to_export:
                            # Берем первую организацию и извлекаем город из адреса
                            first_org = data_to_export[0]
                            full_address = first_org.get('full_address', '')
                            print(f"🔍 Извлекаем город из адреса: '{full_address}'")
                            
                            # Ищем паттерны типа "хутор Бетта", "село Криница", "город Москва", "Геленджик"
                            import re
                            city_patterns = [
                                r'(хутор|село|деревня|посёлок|город|станица|аул|кишлак)\s+([А-Яа-яё]+)',
                                r'([А-Яа-яё]+)\s+(хутор|село|деревня|посёлок|город|станица|аул|кишлак)',
                                r'([А-Яа-яё]+),\s*Россия',  # Паттерн для городов типа "Геленджик, Россия"
                                r'([А-Яа-яё]+),\s*[А-Яа-яё\s]+,\s*[А-Яа-яё\s]+,\s*[А-Яа-яё\s]+',  # Паттерн для "Геленджик, улица..."
                                r'([А-Яа-яё]+),\s*улица',  # Простой паттерн для "Геленджик, улица"
                            ]
                            
                            for i, pattern in enumerate(city_patterns):
                                match = re.search(pattern, full_address, re.IGNORECASE)
                                print(f"🔍 Паттерн {i+1}: '{pattern}' -> {match}")
                                if match:
                                    if 'хутор' in match.group(0).lower() or 'село' in match.group(0).lower():
                                        extracted_city_name = match.group(2) if len(match.groups()) > 1 else match.group(1)
                                        print(f"✅ Найден населенный пункт: '{extracted_city_name}'")
                                        break
                                    elif 'россия' in match.group(0).lower():
                                        # Для паттерна "Город, Россия" берем название города
                                        extracted_city_name = match.group(1)
                                        print(f"✅ Найден город (Россия): '{extracted_city_name}'")
                                        break
                                    elif i == 3:  # Паттерн для "Геленджик, улица..."
                                        extracted_city_name = match.group(1)
                                        print(f"✅ Найден город (общий паттерн): '{extracted_city_name}'")
                                        break
                                    elif i == 4:  # Простой паттерн для "Геленджик, улица"
                                        extracted_city_name = match.group(1)
                                        print(f"✅ Найден город (улица): '{extracted_city_name}'")
                                        break
                        
                        if extracted_city_name:
                            city_name = extracted_city_name.lower()
                            file_source = f"город '{extracted_city_name}' (по координатам)"
                        else:
                            file_source = f"координаты {lat:.4f}, {lon:.4f}, радиус {radius} км"
                else:
                    return jsonify({'error': 'Неверный формат координат'}), 400
            except ValueError:
                return jsonify({'error': 'Ошибка парсинга координат'}), 400
        else:
            return jsonify({'error': 'Не указаны ни город, ни координаты'}), 400
        
        print(f"📊 Запрос на экспорт Excel. Источник: {file_source}, Найдено организаций: {len(data_to_export)}")
        
        if not data_to_export:
            return jsonify({'error': 'Нет данных для экспорта'}), 400
        
        # Создаем новую рабочую книгу
        wb = Workbook()
        ws = wb.active
        ws.title = "Курортные организации"
        
        # Стили для заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Заголовки колонок
        headers = [
            "Название организации",
            "Координаты (широта/долгота)", 
            "ID организации",
            "Полный адрес",
            "Веб-сайт",
            "E-mail",
            "Тип организации",
            "Город поиска"
        ]
        
        # Записываем заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Записываем данные
        for row, org in enumerate(data_to_export, 2):
            # Форматируем координаты
            coords = org.get('coordinates', [])
            coords_str = f"{coords[1]:.6f}, {coords[0]:.6f}" if len(coords) >= 2 else ""
            
            ws.cell(row=row, column=1, value=org.get('name', ''))
            ws.cell(row=row, column=2, value=coords_str)
            ws.cell(row=row, column=3, value=org.get('yandex_id', ''))
            ws.cell(row=row, column=4, value=org.get('full_address', ''))
            ws.cell(row=row, column=5, value=org.get('website', ''))
            ws.cell(row=row, column=6, value=org.get('email', ''))
            ws.cell(row=row, column=7, value=org.get('type', ''))
            ws.cell(row=row, column=8, value=org.get('city', ''))
        
        # Автоподбор ширины колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Создаем файл в памяти
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Генерируем имя файла в зависимости от источника данных
        # Простая транслитерация для безопасного имени файла
        city_mapping = {
            'бэтта': 'betta',
            'москва': 'moscow', 
            'санкт-петербург': 'saint_petersburg',
            'сочи': 'sochi',
            'екатеринбург': 'ekaterinburg',
            'новосибирск': 'novosibirsk',
            'казань': 'kazan',
            'нижний новгород': 'nizhny_novgorod',
            'челябинск': 'chelyabinsk',
            'самара': 'samara',
            'омск': 'omsk',
            'ростов-на-дону': 'rostov_on_don',
            'уфа': 'ufa',
            'красноярск': 'krasnoyarsk',
            'пермь': 'perm',
            'волгоград': 'volgograd',
            'воронеж': 'voronezh',
            'саратов': 'saratov',
            'краснодар': 'krasnodar',
            'тольятти': 'tolyatti',
            'геленджик': 'gelendzhik'
        }
        
        # Используем маппинг или создаем безопасное имя
        safe_city_name = city_mapping.get(city_name.lower(), city_name.lower().replace(' ', '_').replace('-', '_'))
        # Удаляем все нелатинские символы
        safe_city_name = ''.join(c for c in safe_city_name if c.isalpha() or c == '_')
        filename = f"{safe_city_name} + {radius}км.xlsx"
        
        print(f"✅ Excel файл создан: {filename}")
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"❌ Ошибка при создании Excel файла: {str(e)}")
        return jsonify({'error': f'Ошибка при создании Excel файла: {str(e)}'}), 500

@app.route('/api/get_status', methods=['GET'])
def get_status():
    return jsonify({
        'processes': current_processes,
        'organizations_count': len(organizations_data)
    })

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
